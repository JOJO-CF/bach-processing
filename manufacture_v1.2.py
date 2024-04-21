# coding=gbk
from inspect import Parameter
import xml.etree.ElementTree as ET
import pandas as pd
import numpy as np
import pyautogui
import pyperclip
import openpyxl
import win32api
import win32gui
import win32con
import shutil
import time
import math
import os
import re

N = 5  # �������������ʱ������ĸ���,�������ļ��и���,�滻���ɵ��ļ�����,�����������Ĵ���,��õİ�ȫϵ���ĸ���
#����ļ������̿��ƣ��б��е�һ���ֵ��еġ��ǡ��͡��񡱴������͹ر���ع���
process_control = [
    {'�Ƿ������������������Excel���':'��'},
    {'�Ƿ���ӻ����滻����':'��'},
    {'�Ƿ�������������������':'��'},
    {'�Ƿ���ȡ��������Excel���':'��'}
] 
#�б��е�һ���ֵ��еġ��ǡ��͡��񡱴������͹ر���ع���,���ǲ���ͬʱΪ���ǡ��������ֵ��е����ִ�����Ҫ�滻����ӵĲ�������������������깦��δ���
area_coordinates_options = [
    {'�滻��������':'��', '�����������':'��'},
    {'������X':6,'��̬�ֲ���ֵ1': 25, '��̬�ֲ�����1': 56.25,'��̬�ֲ���ֵ2': 5, '��̬�ֲ�����2': 2.25,'��̬�ֲ���ֵ3': 5, '��̬�ֲ�����3': 2.25,'��̬�ֲ���ֵ4': 5, '��̬�ֲ�����4': 2.25,'��̬�ֲ���ֵ5': 5, '��̬�ֲ�����5': 2.25,'��̬�ֲ���ֵ6': 5, '��̬�ֲ�����6': 2.25, 'LeftSideLeftPt X': 2, 'LeftSideRightPt X': 3, 'RightSideLeftPt X': 4, 'RightSideRightPt X': 5},
    {'������Y':6,'��̬�ֲ���ֵ1': 5, '��̬�ֲ�����1': 2.25,'��̬�ֲ���ֵ2': 5, '��̬�ֲ�����2': 2.25,'��̬�ֲ���ֵ3': 5, '��̬�ֲ�����3': 2.25,'��̬�ֲ���ֵ4': 5, '��̬�ֲ�����4': 2.25,'��̬�ֲ���ֵ5': 5, '��̬�ֲ�����5': 2.25,'��̬�ֲ���ֵ6': 5, '��̬�ֲ�����6': 2.25, 'LeftSideLeftPt Y': 2, 'LeftSideRightPt Y': 3, 'RightSideLeftPt Y': 4, 'RightSideRightPt Y': 5},
    {'��������X':4, 'LeftSideLeftPt X': 2, 'LeftSideRightPt X': 3, 'RightSideLeftPt X': 4, 'RightSideRightPt X': 5},
    {'��������Y':4, 'LeftSideLeftPt Y': 2, 'LeftSideRightPt Y': 3, 'RightSideLeftPt Y': 4, 'RightSideRightPt Y': 5}
]   
#�б��е�һ���ֵ��еġ��ǡ��͡��񡱴������͹ر���ع���,���ǲ���ͬʱΪ���ǡ��������ֵ��в����������ִ�����Ҫ�滻����ӵĲ�����������������������δ���
soil_parameter_options = [
    {'�滻�������':'��', '����������':'��'},
    {'��Ħ����':1,'��̬�ֲ���ֵ1': 25, '��̬�ֲ�����1': 56.25},
    {'ճ����':1,'��̬�ֲ���ֵ1': 5, '��̬�ֲ�����1': 2.25},
    {'�ض�':1,'��̬�ֲ���ֵ1': 5, '��̬�ֲ�����1': 2.25}
]   
#�б��е�һ���ֵ��еġ��ǡ��͡��񡱴������͹ر���ع���,���ǲ���ͬʱΪ���ǡ����仯������������3��ԭ��
saturation_line_options = [
    {'�滻������':'��', '��ӽ�����':'��'},
    {'�����ߺ�����':['6', '12', '18', '24'],'����������������':[6, 7, 9, 13],'����������������':[6, 11, 15, 17]},
    {'������Ywn':1, '��̬�ֲ���ֵ1':0.5, '��̬�ֲ�����1':0.5 / 3, '�㶨�ֲ�0':0.5, '�޷ֲ�0':0}
]
#��������������ز���
���������·�� = 'C:\\Program Files (x86)\\GEO-SLOPE\\GeoStudio 9\\Bin\\GeoStudio.exe'  #·���ַ����б������˫��б��
����10ƥ����ڵĴ��ڱ��� = []                                                                   #ʢ������10ƥ��Ĵ��ڱ���
hwnd_title = {}                                                                               #ʢ�Ż�ȡ�Ĵ��ڱ���
                                                                                              # "batch_instruction_set1"ָ��λ�����������ִ�д�(����check_blank_calculate��)
batch_instruction_set2 = [
    {'ָ��1':3,'����':'alt,w,down,down,down,down,enter,space','ѭ������':1},
    {'ָ��2':6,'����':8,'ѭ������':1}
]                                                                                             #������ڴ��Ĳ���
                                                                                              # "batch_instruction_set3"ָ��λ�����������ִ�д�(����check_blank_calculate��)
batch_instruction_set4 = [
    {'ָ��1':8,'����':'taskkill /F /IM GeoStudio.exe','ѭ������':1}
]                                                                                             #�ر����������

#ͨ���������������Ҫ�޸ĵĲ���
key_map = {
    "0": 96, "1": 97, "2": 98, "3": 99, "4": 100, "5": 101, "6": 102, "7": 103, "8": 104, "9": 105, 
    '*': 106, '+': 107, '-': 109, '.': 110, '/': 111,
    'F1': 112, 'F2': 113, 'F3': 114, 'F4': 115, 'F5': 116, 'F6': 117, 'F7': 118, 'F8': 119,
    'F9': 120, 'F10': 121, 'F11': 122, 'F12': 123, 'F13': 124, 'F14': 125, 'F15': 126, 'F16': 127,
    "A": 65, "B": 66, "C": 67, "D": 68, "E": 69, "F": 70, "G": 71, "H": 72, "I": 73, "J": 74,
    "K": 75, "L": 76, "M": 77, "N": 78, "O": 79, "P": 80, "Q": 81, "R": 82, "S": 83, "T": 84,
    "U": 85, "V": 86, "W": 87, "X": 88, "Y": 89, "Z": 90,
    'BACKSPACE': 8, 'TAB': 9, 'TABLE': 9, 'CLEAR': 12, 'ENTER': 13, 'SHIFT': 16, 'CTRL': 17,
    'CONTROL': 17, 'ALT': 18, 'ALTER': 18, 'PAUSE': 19, 'BREAK': 19, 'CAPSLK': 20, 'CAPSLOCK': 20, 
    'ESC': 27,'SPACE': 32, 'SPACEBAR': 32, 'PGUP': 33, 'PAGEUP': 33, 'PGDN': 34,'PAGEDOWN': 34, 
    'END': 35, 'HOME': 36,'LEFT': 37, 'UP': 38, 'RIGHT': 39, 'DOWN': 40, 'SELECT': 41,'PRTSC': 42, 
    'PRINTSCREEN': 42, 'SYSRQ': 42,'SYSTEMREQUEST': 42, 'EXECUTE': 43, 'SNAPSHOT': 44,'INSERT': 45, 
    'DELETE': 46, 'HELP': 47, 'WIN': 91,'WINDOWS': 91, 'NMLK': 144,'NUMLK': 144, 'NUMLOCK': 144, 
    'SCRLK': 145, '[': 219, ']': 221, 
    '������':175, '������':174, 'ֹͣ':179, '����':173, '�����':172, '�ʼ�':180, '����':170
    }                                                                                          #win32api����ֵ��Ӧ���̰��������ΰ������ּ��̰��������ܼ���������ĸ�����������Ƽ���������ý�������
work_dir = os.path.dirname(os.path.abspath(__file__))                                          # ��ȡ�����ļ���·��
original_file_name = 'ģ��'                                                                    # �������滻����ԭʼ�ļ�����
original_file_name_type = 'ģ��.xml'                                                           # �������滻���ݴ���׺ԭʼ�ļ�����
doc_name1 = 'ģ�Ͳ�����'
parameter_options = []
main_folder_name1 = '������Ŀ���ļ���'
for dic in area_coordinates_options + soil_parameter_options + saturation_line_options:
    for k,v in dic.items():
        if v == '��':
            parameter_options.append(k)
            doc_name1 += ',' + k
            main_folder_name1 += ',' + k
doc_name1 += '.xlsx'

doc_name1 = doc_name1.partition(",")[0] + doc_name1.partition(",")[2]                            # �洢���������������ݵ��ļ���
main_folder_name1 = main_folder_name1.partition(",")[0] + main_folder_name1.partition(",")[2]    # ���������������������ļ�����

def main():
    if process_control[0]['�Ƿ������������������Excel���'] == '��':
        generate_random_numbers(doc_name1)                                                # ����������������ȡ׼���õ����ݿɽ�����ע��
    
    if process_control[1]['�Ƿ���ӻ����滻����'] == '��':
        rd_data = []                                                                          # ����ʢ��Excel�ļ������ݵ�����
        io = work_dir + '\\' + doc_name1                                          
                      # ΪExcel�ļ�����λ��·������Excel�ļ�������ʽ��׺
        rd_data.append(pd.read_excel(io))                                                     # ��ȡ���ⲿ��õĴ�����Excel�е�����
        doc_col_name = list(rd_data[0])                                                       # ��ȡExcel�ļ�������
        parameters_name = doc_col_name[1:]                                                    # ��ȡExcel�ļ��ĸ�������
        path = work_dir + '\\' + original_file_name_type                                      # Ϊԭʼxml�ļ�����λ��·������ԭʼxml�ļ�������ʽ��׺

        if area_coordinates_options[0]['�����������'] == '��' or soil_parameter_options[0]['����������'] == '��' or saturation_line_options[0]['��ӽ�����'] == '��':
            replace_add_data(main_folder_name1,rd_data[0],parameters_name)                   # �����滻�����������ӽ�����
        else:
            replace_data(path, main_folder_name1, rd_data[0], 0, parameters_name)         # �����滻��������
            add_data(path, main_folder_name1,rd_data[0],0)                                   # �������ˮѹ��
        move_file(main_folder_name1)                                                      # �����ƶ��ļ�����Ӧ���ļ���
    
    if process_control[2]['�Ƿ�������������������'] == '��':
        check_blank_calculate(main_folder_name1)                                          # �������������������
    
    if process_control[3]['�Ƿ���ȡ��������Excel���'] == '��':
        get_data(doc_name1, main_folder_name1)                                        # ��ȡ��ȫϵ���Լ��������

 # ���ݸ��ֲַ���������
def generate_random_numbers(d_name):
    soil_parameters = {}
    
    if soil_parameter_options[0]['�滻�������'] == '��' or soil_parameter_options[0]['����������'] == '��':
        #����������ز���
        for i in soil_parameter_options:
            if '��Ħ����' in i:
                parameter_name = '��Ħ����'
            elif 'ճ����' in i:
                parameter_name = 'ճ����'
            elif '�ض�' in i:
                parameter_name = '�ض�'
            else:
                continue
            for j in range(i[parameter_name]):
                for k in i:
                    str_num = str(j + 1)
                    if '��̬�ֲ���ֵ' + str_num == k:
                        R = np.random.normal(i['��̬�ֲ���ֵ' + str_num], i['��̬�ֲ�����' + str_num], N)
                        soil_parameters[parameter_name + str_num] = R
                    elif '������̬�ֲ���ֵ' + str_num == k:
                        xm = math.log((i['��̬�ֲ���ֵ' + str_num] ** 2) / math.sqrt(i['��̬�ֲ�����' + str_num] + (i['��̬�ֲ���ֵ' + str_num] ** 2)))
                        xd = math.sqrt(math.log(i['��̬�ֲ�����' + str_num] / (i['��̬�ֲ���ֵ' + str_num] ** 2) + 1))
                        R = np.random.lognormal(xm, xd, N)
                        soil_parameters[parameter_name + str_num] = R
    
    if area_coordinates_options[0]['�滻��������'] == '��' or area_coordinates_options[0]['�����������'] == '��':
        #������������
        for i in area_coordinates_options:
            if '������X' in i:
                parameter_name = '������X'
            elif '������Y' in i:
                parameter_name = '������Y'
            else:
                continue
            for j in range(i[parameter_name]):
                for k in i:
                    str_num = str(j + 1)
                    if '��̬�ֲ���ֵ' + str_num == k:
                        R = np.random.normal(i['��̬�ֲ���ֵ' + str_num], i['��̬�ֲ�����' + str_num], N)
                        soil_parameters[parameter_name + str_num] = R

                    elif '������̬�ֲ���ֵ' + str_num == k:
                        xm = math.log((i['��̬�ֲ���ֵ' + str_num] ** 2) / math.sqrt(i['��̬�ֲ�����' + str_num] + (i['��̬�ֲ���ֵ' + str_num] ** 2)))
                        xd = math.sqrt(math.log(i['��̬�ֲ�����' + str_num] / (i['��̬�ֲ���ֵ' + str_num] ** 2) + 1))
                        R = np.random.lognormal(xm, xd, N)
                        soil_parameters[parameter_name + str_num] = R

    if saturation_line_options[0]['�滻������'] == '��' or saturation_line_options[0]['��ӽ�����'] == '��':
        #���ɱ仯��������ز���
        for i in saturation_line_options:
            if '������Ywn' in i:
                parameter_name = '������Ywn'
            else:
                continue
            for j in range(i[parameter_name]):
                for k in i:
                    str_num = str(j + 1)
                    if '��̬�ֲ���ֵ' + str_num == k:
                        R = np.random.normal(i['��̬�ֲ���ֵ' + str_num], i['��̬�ֲ�����' + str_num], N)
                        soil_parameters[parameter_name + str_num] = R
                    elif '�㶨�ֲ�' + str_num == k:
                        R = [i['�㶨�ֲ�' + str_num] for m in R]
                        soil_parameters[parameter_name + str_num] = R
                    elif '�޷ֲ�' + str_num == k:
                        R = [i['�޷ֲ�' + str_num] for m in R]
                        soil_parameters[parameter_name + str_num] = R

    # �������
    sp = pd.DataFrame(soil_parameters)
    sp.index = sp.index + 1
    if os.path.exists(work_dir + '\\' + d_name):
        print('�Ѵ��ڸ��ļ�')
    else:
        sp.to_excel(work_dir + '\\' + d_name, sheet_name='ԭʼ����', index=1, index_label='ģ��')

# ���������ļ���
def make_dir(m_folder_name):
    path = work_dir + '\\' + m_folder_name
    if not os.path.exists(path):
        os.mkdir(path)
    for i in range(N):
        path = work_dir + '\\' + m_folder_name
        path = path + "\\" + str(i + 1) + '_Runs'
        if not os.path.exists(path):
            os.mkdir(path)


# ���滻�����ļ��ƶ����½����ļ�����
def move_file(m_folder_name):
    cur_file_dir = work_dir + '\\' + m_folder_name
    for i in range(N):
        des_file_dir = work_dir + '\\' + m_folder_name + '\\' + str(i + 1) + '_Runs'
        file_dir = cur_file_dir + '\\' + original_file_name + str(i + 1) + '.xml'
        shutil.move(file_dir, des_file_dir)


# ������ǩ
def prettyXml(element, indent, newline, level=0):  # elemntΪ��������Elment�࣬����indent����������newline���ڻ���
    if element:  # �ж�element�Ƿ�����Ԫ��
        if element.text == None or element.text.isspace():  # ���element��textû������
            element.text = newline + indent * (level + 1)
        else:
            element.text = newline + indent * (level + 1) + element.text.strip() + newline + indent * (level + 1)
    # else:                                                       # �˴����������ע��ȥ����Element��textҲ������һ��
    # element.text = newline + indent * (level + 1) + element.text.strip() + newline + indent * level
    temp = list(element)  # ��elemntת��list
    for subelement in temp:
        if temp.index(subelement) < (len(temp) - 1):  # �������list�����һ��Ԫ�أ�˵����һ������ͬ����Ԫ�ص���ʼ������Ӧһ��
            subelement.tail = newline + indent * (level + 1)
        else:  # �����list�����һ��Ԫ�أ� ˵����һ����ĸԪ�صĽ���������Ӧ����һ��
            subelement.tail = newline + indent * level
        prettyXml(subelement, indent, newline, level=level + 1)  # ����Ԫ�ؽ��еݹ����

# �����滻����
def replace_data(file_path, m_folder_name, rd_data, j, parameters_name):
    # tree = ET.parse(work_dir + '\\' + original_file_name_type)
    #�仯j������ͬʱ������Ӻ��滻����ʱӦ��
    if file_path == work_dir + '\\' + original_file_name_type:
        n = range(0, N)
    else:
        n = range(j, j + 1)
    
    tree = ET.parse(file_path)
    root = tree.getroot()                                  # ��ȡXML�ļ����ڵ�
    Geometries = root.find('Geometries')                   # ��ȡ�ӽڵ�Geometries
    Materials = root.find('Materials')                     # ��ȡ�ӽڵ�Materials
    StabilityItems = root.find('StabilityItems')           # ��ȡ�ӽڵ�StabilityItems
    d = [round(saturation_line_options[1]['����������������'][k] - saturation_line_options[1]['����������������'][k], 3) for k in range(len(saturation_line_options[1]['����������������']))]  # ��ˮλ�߸���ļ��

    for i in n:
        if area_coordinates_options[0]['�滻��������'] == '��' and area_coordinates_options[0]['�����������'] == '��':            
            for Geometry in Geometries.findall('Geometry'):
                Points = Geometry.find('Points')
                for parameter_name in parameters_name:
                    for Point in Points.findall('Point'):
                        if '������' in parameter_name:
                            if '1' in parameter_name and '1' in Point.attrib['ID']:
                                Point.attrib['X'] = str(rd_data[parameter_name][i])
                            elif '2' in parameter_name and '2' in Point.attrib['ID']:
                                Point.attrib['X'] = str(rd_data[parameter_name][i])
                            elif '3' in parameter_name and '3' in Point.attrib['ID']:
                                Point.attrib['X'] = str(rd_data[parameter_name][i])
                            elif '4' in parameter_name and '4' in Point.attrib['ID']:
                                Point.attrib['X'] = str(rd_data[parameter_name][i])
                            elif '5' in parameter_name and '5' in Point.attrib['ID']:
                                Point.attrib['X'] = str(rd_data[parameter_name][i])
                            elif '6' in parameter_name and '6' in Point.attrib['ID']:
                                Point.attrib['X'] = str(rd_data[parameter_name][i])
                        if '������' in parameter_name:
                            if '1' in parameter_name and '1' in Point.attrib['ID']:
                                Point.attrib['Y'] = str(rd_data[parameter_name][i])
                            elif '2' in parameter_name and '2' in Point.attrib['ID']:
                                Point.attrib['Y'] = str(rd_data[parameter_name][i])
                            elif '3' in parameter_name and '3' in Point.attrib['ID']:
                                Point.attrib['Y'] = str(rd_data[parameter_name][i])
                            elif '4' in parameter_name and '4' in Point.attrib['ID']:
                                Point.attrib['Y'] = str(rd_data[parameter_name][i])
                            elif '5' in parameter_name and '5' in Point.attrib['ID']:
                                Point.attrib['Y'] = str(rd_data[parameter_name][i])
                            elif '6' in parameter_name and '6' in Point.attrib['ID']:
                                Point.attrib['Y'] = str(rd_data[parameter_name][i])
            for StabilityItem in StabilityItems.findall('StabilityItem'):
                Entry = StabilityItem.find('Entry')
                SlipSurface = Entry.find('SlipSurface')
                EntryExit = SlipSurface.find('EntryExit')
                LeftSideLeftPt = EntryExit.find('LeftSideLeftPt')
                LeftSideRightPt = EntryExit.find('LeftSideRightPt')
                RightSideLeftPt = EntryExit.find('RightSideLeftPt')
                RightSideRightPt = EntryExit.find('RightSideRightPt')
                for parameter_name in parameters_name:
                    if 'LeftSideLeftPt' in parameter_name and 'X' in parameter_name:
                        LeftSideLeftPt.attrib['X'] = str(rd_data[parameter_name][i])
                    elif 'LeftSideLeftPt' in parameter_name and 'Y' in parameter_name:
                        LeftSideLeftPt.attrib['Y'] = str(rd_data[parameter_name][i])
                    elif 'LeftSideRightPt' in parameter_name and 'X' in parameter_name:
                        LeftSideRightPt.attrib['X'] = str(rd_data[parameter_name][i])
                    elif 'LeftSideRightPt' in parameter_name and 'Y' in parameter_name:
                        LeftSideRightPt.attrib['Y'] = str(rd_data[parameter_name][i])
                    elif 'RightSideLeftPt' in parameter_name and 'X' in parameter_name:
                        RightSideLeftPt.attrib['X'] = str(rd_data[parameter_name][i])
                    elif 'RightSideLeftPt' in parameter_name and 'Y' in parameter_name:
                        RightSideLeftPt.attrib['Y'] = str(rd_data[parameter_name][i])
                    elif 'RightSideRightPt' in parameter_name and 'X' in parameter_name:
                        RightSideRightPt.attrib['X'] = str(rd_data[parameter_name][i])
                    elif 'RightSideRightPt' in parameter_name and 'Y' in parameter_name:
                        RightSideRightPt.attrib['Y'] = str(rd_data[parameter_name][i])
        
        if soil_parameter_options[0]['�滻�������'] == '��' and soil_parameter_options[0]['����������'] == '��':
            for Material in Materials.findall('Material'):
                ID = Material.find('ID').text
                StressStrain = Material.find('StressStrain')
                for parameter_name in parameters_name:
                    if '��Ħ����' + ID in parameter_name:
                        StressStrain.find('PhiPrime').text = str(rd_data[parameter_name][i])
                    elif 'ճ����' + ID in parameter_name:
                        StressStrain.find('CohesionPrime').text = str(rd_data[parameter_name][i])
                    elif '�ض�' + ID in parameter_name:
                        StressStrain.find('UnitWeight').text = str(rd_data[parameter_name][i])

        if saturation_line_options[0]['�滻������'] == '��' and saturation_line_options[0]['��ӽ�����'] == '��':
            for StabilityItem in StabilityItems.findall('StabilityItem'):
                Entry = StabilityItem.find('Entry')
                DataPoints = Entry.find('DataPoints')
                DataPoint = DataPoints.findall('DataPoint')
                for k in DataPoint:
                    k.attrib['X'] = saturation_line_options[1]['�����ߺ�����'][int(k.attrib['Number']) - 1]
                    k.attrib['Y'] = str(
                        saturation_line_options[1]['����������������'][i][int(k.attrib['Number']) - 1] + d[int(k.attrib['Number']) - 1] * replace_data['Ywn'][i])
        
        location = work_dir + '\\' + m_folder_name
        if not os.path.exists(location):
            make_dir(m_folder_name)
            tree.write(location + '/' + original_file_name + str(i + 1) + '.xml', encoding='utf-8',
                    xml_declaration=True)
        else:
            tree.write(location + '/' + original_file_name + str(i + 1) + '.xml', encoding='utf-8',
                    xml_declaration=True)

# ������ӽ�����
def add_data(file_path, m_folder_name, rd_data, j):
    if file_path == work_dir + '\\' + original_file_name_type:
        n = range(0, N)
    else:
        n = range(j, j + 1)
    d = [round(saturation_line_options[1]['����������������'][k] - saturation_line_options[1]['����������������'][k], 3) for k in range(len(saturation_line_options[1]['����������������']))]  # ��ˮλ�߸���ļ��
    # tree = ET.parse(work_dir + '\\' + original_file_name_type)
    tree = ET.parse(file_path)
    root = tree.getroot()                         # ��ȡXML�ļ����ڵ�
    StabilityItems = root.find('StabilityItems')  # ��ȡ�ӽڵ�StabilityItems
    WaterItems = root.find('WaterItems')          #��ȡ�ӽڵ�
    for i in n:
        if saturation_line_options[0]['�滻������'] == '��' and saturation_line_options[0]['��ӽ�����'] == '��':
            for StabilityItem in StabilityItems.findall('StabilityItem'):
                Entry1 = StabilityItem.find('Entry')
                SubElement_Entry0 = ET.SubElement(Entry1, 'DataPoints',
                                  attrib={'Len': str(len(saturation_line_options[1]['�����ߺ�����']))})                # ��Ӻ�attrib�ı�ǩ��atrib����ӵ����ֵ��ʽ��
                for k in range(len(saturation_line_options[1]['�����ߺ�����'])):
                    ET.SubElement(SubElement_Entry0, 'DataPoint',
                                attrib={'Number': str(k + 1), 'X': saturation_line_options[1]['�����ߺ�����'][k], 'Y': str(saturation_line_options[1]['����������������'][k] + d[k] * rd_data['������Ywn1'][i])})                
                
                SubElement_Entry1 = ET.SubElement(Entry1, 'PiezometricLines',
                                  attrib={'Len': '1'})                        # ��Ӻ�attrib�ı�ǩ��atrib����ӵ����ֵ��ʽ��
                SubElement_Entry1_PiezometricLines = ET.SubElement(SubElement_Entry1, 'PiezometricLine',
                                                                attrib={})    # ��Ӻ�attrib�ı�ǩ��atrib����ӵ����ֵ��ʽ��
                SubElement_PiezometricLine0_ID = ET.SubElement(SubElement_Entry1_PiezometricLines, 'ID')
                SubElement_PiezometricLine0_ID.text = '1'                     # ����text��ע�ⲻ��ֱ����int���͵�
                SubElement_PiezometricLine0_DataPoints = ET.SubElement(SubElement_Entry1_PiezometricLines, 'DataPoints',
                                                                    attrib={'Len': str(len(saturation_line_options[1]['�����ߺ�����']))})
                for k in reversed(range(len(saturation_line_options[1]['�����ߺ�����']))):
                    SubElement_PiezometricLine0_DataPoint0 = ET.SubElement(SubElement_PiezometricLine0_DataPoints,
                                                                        'DataPoint')
                    SubElement_PiezometricLine0_DataPoint0.text = str(k + 1)  # ����text��ע�ⲻ��ֱ����int���͵�
                SubElement_Entry2 = ET.SubElement(Entry1, 'MaterialUsesPiezs',
                                                attrib={'Len': '1'})          # ��Ӻ�attrib�ı�ǩ��atrib����ӵ����ֵ��ʽ��
                ET.SubElement(SubElement_Entry2, 'MaterialUsesPiez', attrib={'ID': '1', 'UsesID': '1'})
                prettyXml(Entry1, '    ', '\n')                                # ������ǩ
                
            for WaterItem in WaterItems.findall('WaterItem'):
                Entry2 = WaterItem.find('Entry')
                SubElement_Entry3 = ET.SubElement(Entry2, 'ResultInputInfo', attrib={})  # ��Ӻ�attrib�ı�ǩ��atrib����ӵ����ֵ��ʽ��
                SubElement_ResultInputInfo_Option = ET.SubElement(SubElement_Entry3,
                                                                'Option')                # ��Ӻ�attrib�ı�ǩ��atrib����ӵ����ֵ��ʽ��
                SubElement_ResultInputInfo_Option.text = 'PiezoLine'                     # ����text��ע�ⲻ��ֱ����int���͵�
                prettyXml(Entry2, '    ', '\n')
                
            location = work_dir + '\\' + m_folder_name
            if not os.path.exists(location):
                make_dir(m_folder_name)
                tree.write(location + '/' + original_file_name + str(i + 1) + '.xml', encoding='utf-8',
                        xml_declaration=True)
            else:
                tree.write(location + '/' + original_file_name + str(i + 1) + '.xml', encoding='utf-8',
                        xml_declaration=True)


def replace_add_data(m_folder_name, rd_data, parameters_name):
    files_folder_path = work_dir + '\\' + m_folder_name
    original_file_path = work_dir + '\\' + original_file_name_type
    if area_coordinates_options[0]['�滻��������'] == '��' and area_coordinates_options[0]['�����������'] == '��':
        pass
    if soil_parameter_options[0]['�滻�������'] == '��' and soil_parameter_options[0]['����������'] == '��':
        pass
    if saturation_line_options[0]['�滻������'] == '��' and saturation_line_options[0]['��ӽ�����'] == '��':
        for i in range(N):
            files_path = files_folder_path + '\\' + original_file_name + str(i + 1) + '.xml'
            if os.path.exists(files_path):
                tree = ET.parse(files_path)
                root = tree.getroot()  # ��ȡXML�ļ����ڵ�
                StabilityItems = root.find('StabilityItems')
                StabilityItem = StabilityItems.find('StabilityItem')
                Entry = StabilityItem.find('Entry')
                DataPoints = Entry.find('DataPoints')
                if DataPoints:
                    replace_data(files_path, m_folder_name, rd_data, i, parameters_name)
            else:
                add_data(original_file_path, m_folder_name, rd_data, 0)
                replace_data(files_path, m_folder_name, rd_data, i, parameters_name)

#�����ز�������BվUP�� �����˾ͺ�ˮ �ṩ
def mouseClick(clickTimes,lOrR,img):
    while True:
        location=pyautogui.locateCenterOnScreen(img,confidence=0.9)
        if location is not None:
            pyautogui.click(location.x,location.y,clicks=clickTimes,interval=0.2,duration=0.2,button=lOrR)
            break
        print("δ�ҵ�ƥ��ͼƬ,0.1�������")
        time.sleep(0.1)

#��������������������,��Bվup�� ������˫ �ṩ��
def presskey(hk_g_inputValue):
    keys = hk_g_inputValue.split(',')
    for key in keys:
        if isinstance(key, str):
            win32api.keybd_event(key_map[key.upper()], win32api.MapVirtualKey(key_map[key.upper()], 0), 0, 0)
            win32api.keybd_event(key_map[key.upper()], win32api.MapVirtualKey(key_map[key.upper()], 0), win32con.KEYEVENTF_KEYUP, 0)
            time.sleep(0.1)
    print("ִ���ˣ�",hk_g_inputValue)
    time.sleep(0.1)

#�ж��ȼ���ϸ���������������,��Bվup�� ������˫ �ṩ
def hotkey_get(hk_g_inputValue):
    try:
        keys = hk_g_inputValue.split(',')
        for key in keys:
            if isinstance(key, str):
                win32api.keybd_event(key_map[key.upper()], win32api.MapVirtualKey(key_map[key.upper()], 0), 0, 0)
        for key in keys:
            if isinstance(key, str):
                win32api.keybd_event(key_map[key.upper()], win32api.MapVirtualKey(key_map[key.upper()], 0), win32con.KEYEVENTF_KEYUP, 0)
        print("ִ���ˣ�",hk_g_inputValue)
        time.sleep(0.1)
    except:
        pyperclip.copy(hk_g_inputValue)
        pyautogui.hotkey('ctrl', 'v')

#��ȡ���ڱ���
def get_all_hwnd(hwnd, mouse):
    if (win32gui.IsWindow(hwnd) and win32gui.IsWindowEnabled(hwnd) and win32gui.IsWindowVisible(hwnd)):
        hwnd_title.update({hwnd: win32gui.GetWindowText(hwnd)})

def instruction_set_execution(bis):
    i = 0
    while i < len(bis):
        cmdType = bis[i]['ָ��'+ str(i+1)]     #��ȡָ���
        cmdContent = bis[i]['����']            #��ȡָ������
        cycles = bis[i]['ѭ������']                #��ȡָ��ѭ������
        if cmdType == 1:           
            for j in range(cycles):
                if '�������' in cmdContent:
                    mouseClick(1,"left", cmdContent)
                    print("�������", cmdContent)
                elif '˫�����' in cmdContent:
                    mouseClick(2,"left", cmdContent)
                    print("˫�����", cmdContent)    
                elif '�����Ҽ�' in cmdContent:
                    mouseClick(1,"right", cmdContent)
                    print("�����Ҽ�", cmdContent) 
        #2�����ֲ���
        elif cmdType == 2:
            for j in range(cycles):
                pyautogui.scroll(cmdContent)
                print("���ֻ���",int(cmdContent),"����")     
        #3���̰���
        elif cmdType == 3:
            for j in range(cycles):
                presskey(cmdContent)
                time.sleep(0.5)
        #4�����ȼ����
        elif cmdType == 4:
            for j in range(cycles):
                hotkey_get(cmdContent)
                time.sleep(0.5)
        #5�����ַ���
        elif cmdType == 5:
            for j in range(cycles):
                pyperclip.copy(cmdContent)
                pyautogui.hotkey('ctrl','v')
                print("����:",cmdContent) 
                time.sleep(0.5)                                       
        #6�ȴ�
        elif cmdType == 6:
            for j in range(cycles):
                time.sleep(cmdContent)
                print("�ȴ�",cmdContent,"��")
        #7ճ����ǰʱ��
        elif cmdType == 7:      
            for j in range(cycles):
                localtime = time.strftime("%Y-%m-%d %H��%M��%S", time.localtime())  #���ñ�����ǰʱ�䡣
                pyperclip.copy(localtime)
                pyautogui.hotkey('ctrl','v')
                print("ճ���˱���ʱ��:",localtime)
                time.sleep(0.5)
        #8ϵͳ���
        elif cmdType == 8: 
            for j in range(cycles):
                os.system(cmdContent)
                print("����ϵͳ����:",cmdContent)
                time.sleep(0.5) 
        #9����ָ�������ָ���ļ����ߴ�ĳ����
        elif cmdType == 9:
            cmdContent = cmdContent.split(',') 
            for j in range(cycles):
                if len(cmdContent) == 1:
                    win32api.ShellExecute(0, 'open', cmdContent[0], "", "", 1)                                      #ֻ��ָ������1��������ָ����������·��
                elif len(cmdContent) == 2:
                    win32api.ShellExecute(0, 'open', cmdContent[0], cmdContent[1], "", 1)  #����ָ�������Ĭ���ļ�·����ָ���ļ���2��������ָ����������·�����ļ����ƣ�������ʽ��׺��
                elif len(cmdContent) == 3:
                    win32api.ShellExecute(0, 'open', cmdContent[0], cmdContent[1], cmdContent[2], 1)  #����ָ�������ָ���ļ���ָ���ļ���3��������ָ����������·�����ļ����ƣ�������ʽ��׺�����ļ������ļ���·��
                print("����ϵͳ����:",cmdContent)
                time.sleep(0.5) 
        #10 ������������д򿪴��ڵı��⣬��ѡ�ƥ�䴰�ڱ���
        elif cmdType == 10:
            cmdContent = cmdContent.split(',')
            for j in range(cycles): 
                win32gui.EnumWindows(get_all_hwnd, 0)
                print('������ڴ���:', hwnd_title)
                for k in cmdContent:
                    for m in hwnd_title.values():
                        if k == m:
                            ����10ƥ����ڵĴ��ڱ���.append(k)
                            print('ƥ�䴰��:', k, '����')
                time.sleep(0.5) 
        i += 1

# �����ж�λ�ò�������߽�������������GeoStudio���м���
def check_blank_calculate(m_folder_name):
    file_dir = work_dir + '\\' + m_folder_name
    # print(file_dir)
    file_handle = open(work_dir + '\\' + '����λ��.txt', mode='w')
    location = []
    for i in range(N):
        cur_dir1 = file_dir + '\\' + str(i + 1) + '_Runs\\SLOPE&3W Analysis\\001'
        cur_dir2 = file_dir + '\\' + str(i + 1) + '_Runs\\SLOPE&3W ����\\001'
        # print(cur_dir)
        if os.path.exists(cur_dir1):
            cur_dir = cur_dir1
        elif os.path.exists(cur_dir2):
            cur_dir = cur_dir2
        if not os.path.exists(cur_dir):
            location.append(i + 1)
            file_handle.write(str(i + 1) + '\n')

    for i in location:
        Path = work_dir + '\\' + m_folder_name + '\\' + str(i) + '_Runs\\'
        file = original_file_name + str(i) + '.xml'
        print(Path, file)
        # ����ָ�������ָ���ļ����µ�ָ���ļ����ȴ���ƥ��������ڣ�ƥ��Ĵ��ڳ��ֺ���п�ݼ�����
        batch_instruction_set1 = [
            {'ָ��1':9,'����':���������·�� + ',' + file + ',' + Path,'ѭ������':1},
            {'ָ��2':6,'����':10,'ѭ������':1},
            {'ָ��3':10,'����':original_file_name + str(i) + '.xml - GeoStudio 2018 R2 (SLOPE/W Definition)','ѭ������':1}
        ]
        instruction_set_execution(batch_instruction_set1)
        for j in ����10ƥ����ڵĴ��ڱ���:
            if j == original_file_name + str(i) + '.xml - GeoStudio 2018 R2 (SLOPE/W Definition)':
                instruction_set_execution(batch_instruction_set2)

        # �ȴ���ƥ���Ƿ���ڽ�����ڣ���������رճ��򣬷������
        batch_instruction_set3 = [
            {'ָ��1':6,'����':2,'ѭ������':1},
            {'ָ��2':10,'����':original_file_name + str(i) + '.xml - GeoStudio 2018 R2 (SLOPE/W Results)','ѭ������':1},
            {'ָ��3':10,'����':original_file_name + str(i) + '.xml - GeoStudio 2018 R2 (SLOPE/W Definition)','ѭ������':1}
        ]
        instruction_set_execution(batch_instruction_set3)
        for j in ����10ƥ����ڵĴ��ڱ���:
            if j == original_file_name + str(i) + '.xml - GeoStudio 2018 R2 (SLOPE/W Results)':
                instruction_set_execution(batch_instruction_set4)
            elif j == original_file_name + str(i) + '.xml - GeoStudio 2018 R2 (SLOPE/W Definition)':
                instruction_set_execution(batch_instruction_set4)
            else:
                continue

# ��ȡ��ȫϵ���ͻ�����������
def get_data(d_name, m_main_folder_name):
    file_dir = work_dir + '\\' + m_main_folder_name
    # print(file_dir)
    fs_data = ['��ȫϵ��']
    sv_data = ['�������']
    for i in range(N):
        cur_dir1 = file_dir + '\\' + str(i + 1) + '_Runs\\SLOPE&3W Analysis\\001'
        cur_dir2 = file_dir + '\\' + str(i + 1) + '_Runs\\SLOPE&3W ����\\001'
        # print(cur_dir)
        if os.path.exists(cur_dir1):
            cur_dir = cur_dir1
        elif os.path.exists(cur_dir2):
            cur_dir = cur_dir2
        msp_number = 1
        for root, dirs, files in os.walk(cur_dir):
            for file in files:
                if 'lambdafos_' in file and file.endswith('.csv'):
                    msp_number = int(re.findall('\d+', file)[0]) - 1
                    cur_files1 = root + '\\' + file
                    data1 = pd.read_csv(cur_files1)
                    if len(data1) == 0:
                        fs_data.append(0)
                    else:
                        fs_data.append(data1['FOSByMoment'][len(data1) - 1])
                if 'slip_surface' in file and file.endswith('.csv'):
                    cur_files2 = root + '\\' + file
                    data2 = pd.read_csv(cur_files2)
                    sv_data.append(data2['SlipVolume'][msp_number])
    wb = openpyxl.load_workbook(work_dir + '\\' + d_name)  # ����һ���Ѵ��ڵ�wookbook����
    wb1 = wb.active  # ����sheet
    tcl_number = 1
    if area_coordinates_options[0]['�滻��������'] == '��' or area_coordinates_options[0]['�����������'] == '��':
        tcl_number += area_coordinates_options[1]['������X'] + area_coordinates_options[2]['������Y'] + area_coordinates_options[3]['��������X'] + area_coordinates_options[4]['��������Y']
    elif soil_parameter_options[0]['�滻�������'] == '��' or soil_parameter_options[0]['����������'] == '��':
        tcl_number += soil_parameter_options[1]['��Ħ����'] + soil_parameter_options[2]['ճ����'] + soil_parameter_options[3]['�ض�']
    elif saturation_line_options[0]['�滻������'] == '��' or saturation_line_options[0]['��ӽ�����'] == '��':
        tcl_number += 1
    for i in range(len(fs_data)):
        wb1.cell(i + 1, tcl_number).value = fs_data[i]
    for i in range(len(sv_data)):
        wb1.cell(i + 1, tcl_number + 1).value = sv_data[i]
    wb.save(work_dir + '\\' + d_name)  # ����



if __name__ == "__main__":
    main()
